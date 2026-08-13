from __future__ import annotations

import asyncio
import hashlib
import io
import logging
from dataclasses import dataclass
from datetime import date
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, sessionmaker

from app.modules.assets.model import ExternalSourceModel
from app.modules.inventory.persistence_model import (
    InventoryDailyRunModel,
    InventoryDocumentModel,
    InventoryDocumentPageModel,
    InventoryExportModel,
    InventorySettingsModel,
    InventorySourceFileModel,
    InventoryTransactionModel,
    inventory_utcnow,
)
from app.providers.google.auth import get_connection_access_token
from app.providers.google.drive import GoogleDriveClient

logger = logging.getLogger(__name__)
PROTECTED_SHEET = "Báo cáo sử dụng NVL trong ca"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class InventoryExportFailure(ValueError):
    """A safe, stable export failure code suitable for the Inventory API."""


@dataclass(frozen=True)
class ExportResult:
    id: str
    business_date: date
    status: str
    main_drive_file_id: str | None
    backup_drive_file_id: str | None
    content_sha256: str | None
    completed_at: object
    error_code: str | None
    archive_status: str
    archive_error_code: str | None


def _style_fingerprint(cell) -> tuple:
    return (
        str(cell.font), str(cell.fill), str(cell.border), str(cell.alignment),
        cell.number_format, str(cell.protection),
    )


def sheet_fingerprint(sheet) -> str:
    """Fingerprint every Sheet 4 property that export code must preserve."""
    cells = []
    for row in sheet.iter_rows():
        for cell in row:
            cells.append((
                cell.coordinate, cell.value, cell.data_type, _style_fingerprint(cell),
            ))
    rows = sorted((key, dim.height, dim.hidden) for key, dim in sheet.row_dimensions.items())
    columns = sorted((key, dim.width, dim.hidden) for key, dim in sheet.column_dimensions.items())
    payload = repr((
        sheet.title,
        sheet.parent.sheetnames.index(sheet.title),
        sheet.max_row,
        sheet.max_column,
        sheet.calculate_dimension(),
        sorted(str(item) for item in sheet.merged_cells.ranges),
        cells,
        rows,
        columns,
        sheet.freeze_panes,
        sheet.sheet_view.showGridLines,
        sheet.sheet_state,
    )).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


class InventoryExportService:
    """Export a finalized Inventory snapshot without touching Creative data."""

    def __init__(
        self,
        session_factory: sessionmaker[Session],
        *,
        token_resolver: Callable = get_connection_access_token,
        client_factory: Callable = GoogleDriveClient,
        workbook_factory: Callable[[], Workbook] | None = None,
        shadow_mode: bool = False,
    ):
        self.session_factory = session_factory
        self.token_resolver = token_resolver
        self.client_factory = client_factory
        self.workbook_factory = workbook_factory or Workbook
        self.shadow_mode = shadow_mode

    def get(self, tenant_id: str, business_date: date) -> ExportResult | None:
        with self.session_factory() as session:
            row = session.scalar(
                select(InventoryExportModel)
                .where(
                    InventoryExportModel.tenant_id == tenant_id,
                    InventoryExportModel.business_date == business_date,
                )
                .order_by(InventoryExportModel.created_at.desc())
            )
            return self._view(row) if row else None

    def export(self, tenant_id: str, business_date: date, actor_id: str | None = None) -> ExportResult:
        """Create/retry one logical export. A competing request observes it."""
        if self.shadow_mode:
            logger.info("inventory_shadow_mode_blocked tenant_id=%s operation=excel_export", tenant_id)
            raise InventoryExportFailure("inventory_shadow_mode_export_blocked")
        with self.session_factory.begin() as session:
            if session.bind and session.bind.dialect.name == "postgresql":
                session.execute(
                    text("SELECT pg_advisory_xact_lock(hashtext(:key))"),
                    {"key": f"inventory-export:{tenant_id}:{business_date}"},
                )
            run = session.scalar(
                select(InventoryDailyRunModel).where(
                    InventoryDailyRunModel.tenant_id == tenant_id,
                    InventoryDailyRunModel.business_date == business_date,
                )
            )
            if not run or run.status != "finalized":
                raise InventoryExportFailure("inventory_daily_run_not_finalized")
            settings = session.scalar(
                select(InventorySettingsModel).where(InventorySettingsModel.tenant_id == tenant_id)
            )
            if not self._is_configured(settings):
                raise InventoryExportFailure("inventory_export_not_configured")

            key = f"inventory-export:v1:{run.id}:{run.report_version + 1}"
            row = session.scalar(
                select(InventoryExportModel).where(
                    InventoryExportModel.tenant_id == tenant_id,
                    InventoryExportModel.idempotency_key == key,
                )
            )
            if row and row.status == "completed":
                completed_id = row.id
                retry_archive = row.archive_status == "retryable_failure"
                source_id = settings.external_source_id
                if not retry_archive:
                    return self._view(row)
            else:
                completed_id = None
                source_id = None
                retry_archive = False
            if row and row.status == "generating":
                return self._view(row)
            if row is None:
                try:
                    with session.begin_nested():
                        row = InventoryExportModel(
                            tenant_id=tenant_id,
                            daily_run_id=run.id,
                            idempotency_key=key,
                            export_version=run.report_version + 1,
                            export_format="xlsx",
                            period_month=business_date.strftime("%Y-%m"),
                            business_date=business_date,
                            status="generating",
                            archive_status="pending" if settings.archive_enabled else "not_requested",
                            attempt_count=0,
                            created_by=actor_id,
                        )
                        session.add(row)
                        session.flush()
                except IntegrityError:
                    row = session.scalar(
                        select(InventoryExportModel).where(
                            InventoryExportModel.tenant_id == tenant_id,
                            InventoryExportModel.idempotency_key == key,
                        )
                    )
                    return self._view(row)

            transactions = list(session.scalars(
                select(InventoryTransactionModel)
                .where(
                    InventoryTransactionModel.tenant_id == tenant_id,
                    InventoryTransactionModel.business_date == business_date,
                    InventoryTransactionModel.status == "posted",
                )
                .order_by(InventoryTransactionModel.id)
            ))
            try:
                content = self._workbook(business_date, transactions)
            except InventoryExportFailure as exc:
                row.status = "failed"
                row.error_code = str(exc)
                row.error_message = "Workbook validation failed before any Drive operation."
                validation_failure = str(exc)
                export_id = row.id
                content = None
            if content is None:
                pass
            else:
                row.status = "generating"
                row.error_code = None
                row.error_message = None
                row.attempt_count += 1
                row.started_at = inventory_utcnow()
                row.content_sha256 = hashlib.sha256(content).hexdigest()
                row.metadata_json = {
                    **(row.metadata_json or {}),
                    "main_name": self.main_name(business_date),
                    "backup_name": self.backup_name(business_date),
                    "template_version": "inventory-v1",
                }
                export_id = row.id
                source_id = settings.external_source_id
                folders = (settings.excel_folder_id, settings.backup_folder_id)


        if content is None:
            raise InventoryExportFailure(validation_failure)
        if retry_archive:
            self._archive_after_export(tenant_id, completed_id, business_date, source_id)
            with self.session_factory() as session:
                return self._view(session.get(InventoryExportModel, completed_id))
        try:
            main_id, backup_id = asyncio.run(
                self._upload(tenant_id, source_id, folders, content, business_date, export_id)
            )
        except Exception as exc:
            logger.warning("inventory_export_upload_failed", extra={"tenant_id": tenant_id, "export_id": export_id})
            with self.session_factory.begin() as session:
                row = session.get(InventoryExportModel, export_id)
                row.status = "failed"
                row.error_code = "inventory_export_drive_failure"
                row.error_message = "Inventory export upload failed; retry is safe."
            raise InventoryExportFailure("inventory_export_drive_failure") from exc

        with self.session_factory.begin() as session:
            row = session.get(InventoryExportModel, export_id)
            row.status = "completed"
            row.drive_file_id = main_id
            row.backup_drive_file_id = backup_id
            row.completed_at = inventory_utcnow()
            row.metadata_json = {**(row.metadata_json or {}), "backup_drive_file_id": backup_id}

        self._archive_after_export(tenant_id, export_id, business_date, source_id)
        with self.session_factory() as session:
            return self._view(session.get(InventoryExportModel, export_id))

    @staticmethod
    def _is_configured(settings: InventorySettingsModel | None) -> bool:
        return bool(
            settings
            and settings.excel_export_enabled
            and settings.excel_folder_id
            and settings.backup_folder_id
        )

    def _populate_workbook(self, workbook, business_date: date, transactions) -> None:
        daily = workbook.active
        daily.title = "Báo cáo ngày"
        daily.delete_rows(1, daily.max_row)
        daily.append(["Ngày", "Loại", "Kho", "Mã hàng", "Số lượng", "Đơn vị", "Phiếu"])
        for cell in daily[1]:
            cell.font = Font(bold=True)
        for transaction in transactions:
            daily.append([
                business_date.isoformat(), transaction.transaction_type, transaction.location_id,
                transaction.item_id, str(transaction.quantity_base_unit),
                transaction.base_unit_snapshot, transaction.source_document_id,
            ])

    def _workbook(self, business_date: date, transactions) -> bytes:
        workbook = self.workbook_factory()
        if not workbook.sheetnames:
            workbook.create_sheet()
        for name in ("Tổng hợp tháng", "Đối chiếu"):
            if name not in workbook.sheetnames:
                workbook.create_sheet(name)
        if PROTECTED_SHEET not in workbook.sheetnames:
            workbook.create_sheet(PROTECTED_SHEET)
        protected = workbook[PROTECTED_SHEET]
        before = sheet_fingerprint(protected)
        self._populate_workbook(workbook, business_date, transactions)
        if sheet_fingerprint(protected) != before:
            raise InventoryExportFailure("inventory_sheet4_invariant_failed")
        output = io.BytesIO()
        workbook.save(output)
        restored = load_workbook(io.BytesIO(output.getvalue()))
        if restored.sheetnames.index(PROTECTED_SHEET) != 3 or sheet_fingerprint(restored[PROTECTED_SHEET]) != before:
            raise InventoryExportFailure("inventory_sheet4_invariant_failed")
        return output.getvalue()
    async def _upload(
        self,
        tenant_id: str,
        source_id: str,
        folders: tuple[str, str],
        content: bytes,
        business_date: date,
        export_id: str,
    ) -> tuple[str, str]:
        with self.session_factory() as session:
            source = session.scalar(
                select(ExternalSourceModel).where(
                    ExternalSourceModel.tenant_id == tenant_id,
                    ExternalSourceModel.id == source_id,
                    ExternalSourceModel.source_type == "google_drive",
                )
            )
            existing = session.get(InventoryExportModel, export_id)
            connection = (source.source_metadata or {}).get("oauth_connection_id") if source else None
        if not connection:
            raise InventoryExportFailure("inventory_drive_connection_unavailable")
        token = await self.token_resolver(connection)
        async with self.client_factory(token) as drive:
            main_id = existing.drive_file_id
            if not main_id:
                main = await drive.upload_file(folders[0], self.main_name(business_date), XLSX_MIME, content)
                main_id = main.id
                with self.session_factory.begin() as session:
                    row = session.get(InventoryExportModel, export_id)
                    row.drive_file_id = main_id
            backup_id = existing.backup_drive_file_id
            if not backup_id:
                backup = await drive.upload_file(folders[1], self.backup_name(business_date), XLSX_MIME, content)
                backup_id = backup.id
            return main_id, backup_id

    def _archive_after_export(
        self, tenant_id: str, export_id: str, business_date: date, source_id: str
    ) -> None:
        """Move only finalized/reupload documents after both export files exist."""
        with self.session_factory() as session:
            settings = session.scalar(
                select(InventorySettingsModel).where(InventorySettingsModel.tenant_id == tenant_id)
            )
            if not settings or not settings.archive_enabled:
                return
            if not settings.processed_folder_id or not settings.reupload_folder_id:
                self._archive_failure(export_id, "inventory_archive_not_configured")
                return
            rows = list(session.execute(
                select(
                    InventorySourceFileModel.id,
                    InventorySourceFileModel.drive_file_id,
                    InventorySourceFileModel.provider_metadata_json,
                    InventoryDocumentModel.status,
                )
                .join(InventoryDocumentPageModel, InventoryDocumentPageModel.source_file_id == InventorySourceFileModel.id)
                .join(InventoryDocumentModel, InventoryDocumentModel.id == InventoryDocumentPageModel.document_id)
                .where(
                    InventorySourceFileModel.tenant_id == tenant_id,
                    InventorySourceFileModel.external_source_id == source_id,
                    InventoryDocumentModel.tenant_id == tenant_id,
                    InventoryDocumentModel.business_date == business_date,
                    InventoryDocumentModel.status.in_(("finalized", "needs_reupload")),
                )
            ).all())
            connection_source = session.scalar(select(ExternalSourceModel).where(
                ExternalSourceModel.tenant_id == tenant_id,
                ExternalSourceModel.id == source_id,
                ExternalSourceModel.source_type == "google_drive",
            ))
            connection = (connection_source.source_metadata or {}).get("oauth_connection_id") if connection_source else None
            destinations = {
                "finalized": settings.processed_folder_id,
                "needs_reupload": settings.reupload_folder_id,
            }
        if not connection:
            self._archive_failure(export_id, "inventory_drive_connection_unavailable")
            return
        try:
            asyncio.run(self._move_rows(connection, rows, destinations, business_date, tenant_id))
        except Exception:
            logger.warning("inventory_export_archive_failed", extra={"tenant_id": tenant_id, "export_id": export_id})
            self._archive_failure(export_id, "inventory_archive_drive_failure")
            return
        with self.session_factory.begin() as session:
            row = session.get(InventoryExportModel, export_id)
            row.archive_status = "completed"
            row.archive_error_code = None

    async def _move_rows(self, connection: str, rows, destinations: dict[str, str], business_date: date, tenant_id: str) -> None:
        token = await self.token_resolver(connection)
        async with self.client_factory(token) as drive:
            folder_by_status = {}
            for status, parent_id in destinations.items():
                folder_by_status[status] = (await drive.ensure_child_folder(parent_id, business_date.isoformat())).id
            moved: set[str] = set()
            for source_file_id, drive_file_id, metadata, status in rows:
                if source_file_id in moved:
                    continue
                moved.add(source_file_id)
                archive = (metadata or {}).get("inventory_archive") or {}
                target_id = folder_by_status[status]
                if archive.get("destination_folder_id") == target_id:
                    continue
                await drive.move_file(drive_file_id, target_id)
                with self.session_factory.begin() as session:
                    source = session.get(InventorySourceFileModel, source_file_id)
                    source.provider_metadata_json = {
                        **(source.provider_metadata_json or {}),
                        "inventory_archive": {
                            "destination_folder_id": target_id,
                            "business_date": business_date.isoformat(),
                            "document_status": status,
                        },
                    }

    def _archive_failure(self, export_id: str, code: str) -> None:
        with self.session_factory.begin() as session:
            row = session.get(InventoryExportModel, export_id)
            row.archive_status = "retryable_failure"
            row.archive_error_code = code

    @staticmethod
    def main_name(day: date) -> str:
        return f"inventory-{day:%Y-%m}.xlsx"

    @staticmethod
    def backup_name(day: date) -> str:
        return f"inventory-{day.isoformat()}.xlsx"

    @staticmethod
    def _view(row: InventoryExportModel) -> ExportResult:
        return ExportResult(
            row.id, row.business_date, row.status, row.drive_file_id,
            row.backup_drive_file_id, row.content_sha256, row.completed_at,
            row.error_code, row.archive_status, row.archive_error_code,
        )
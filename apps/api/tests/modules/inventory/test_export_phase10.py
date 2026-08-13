from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import sessionmaker

from app.core.database import Base
from app.modules.auth_persistence.model import TenantModel
from app.modules.assets.model import ExternalSourceModel
from app.modules.inventory.exports.service import (
    InventoryExportFailure,
    InventoryExportService,
    PROTECTED_SHEET,
    sheet_fingerprint,
)
from app.modules.inventory.persistence_model import (
    InventoryDailyRunModel,
    InventoryExportModel,
    InventorySettingsModel,
    InventorySourceFileModel,
)

DAY = date(2030, 8, 9)


class ExportPhase10Test(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.engine = create_engine(f"sqlite:///{Path(self.tmp.name) / 'x.db'}")
        event.listen(self.engine, "connect", lambda connection, _: connection.execute("PRAGMA foreign_keys=ON"))
        keep = {
            "tenants", "external_sources", "inventory_settings", "inventory_daily_runs",
            "inventory_exports", "inventory_transactions", "inventory_source_files",
        }
        for table in Base.metadata.sorted_tables:
            if table.name in keep:
                table.create(self.engine)
        self.sessions = sessionmaker(bind=self.engine, expire_on_commit=False)
        with self.sessions.begin() as session:
            session.add(TenantModel(id="t", name="t", slug="t"))
            session.add(ExternalSourceModel(
                id="src", tenant_id="t", source_key="src", source_type="google_drive",
                source_metadata={"oauth_connection_id": "connection"},
            ))
            session.add(InventorySettingsModel(
                tenant_id="t", external_source_id="src", inbox_folder_id="inbox",
                excel_folder_id="excel", backup_folder_id="backup", excel_export_enabled=True,
            ))
            session.add(InventoryDailyRunModel(
                tenant_id="t", business_date=DAY, idempotency_key="day", status="finalized",
            ))
        self.calls = []

        class Drive:
            async def __aenter__(inner):
                return inner

            async def __aexit__(inner, *_):
                return None

            async def upload_file(inner, parent, name, mime, content):
                self.calls.append((parent, name, content))
                return type("Node", (), {"id": f"{parent}-{name}"})()

        async def token(_):
            return "recognizable-access-token"

        self.service = InventoryExportService(
            self.sessions, token_resolver=token, client_factory=lambda _: Drive()
        )

    def tearDown(self):
        self.engine.dispose()
        self.tmp.cleanup()

    def test_finalized_exports_idempotently_with_main_and_backup_metadata(self):
        first = self.service.export("t", DAY, "actor")
        second = self.service.export("t", DAY, "actor")
        self.assertEqual(first.id, second.id)
        self.assertEqual("completed", first.status)
        self.assertEqual(2, len(self.calls))
        self.assertEqual("excel", self.calls[0][0])
        self.assertEqual("backup", self.calls[1][0])
        self.assertEqual("inventory-2030-08.xlsx", self.calls[0][1])
        self.assertEqual("inventory-2030-08-09.xlsx", self.calls[1][1])
        with self.sessions() as session:
            exports = list(session.scalars(select(InventoryExportModel)))
            self.assertEqual(1, len(exports))
            self.assertEqual(first.backup_drive_file_id, exports[0].backup_drive_file_id)
            self.assertIsNotNone(exports[0].content_sha256)

    def test_backup_retry_reuses_uploaded_main_without_duplicate(self):
        calls = []
        class Drive:
            async def __aenter__(inner): return inner
            async def __aexit__(inner, *_): return None
            async def upload_file(inner, parent, name, mime, content):
                calls.append((parent, name))
                if parent == "backup" and calls.count((parent, name)) == 1:
                    raise RuntimeError("temporary Drive outage")
                return type("Node", (), {"id": f"{parent}-{name}"})()
        service = InventoryExportService(
            self.sessions, token_resolver=self.service.token_resolver, client_factory=lambda _: Drive()
        )
        with self.assertRaisesRegex(InventoryExportFailure, "drive_failure"):
            service.export("t", DAY)
        result = service.export("t", DAY)
        self.assertEqual("completed", result.status)
        self.assertEqual(1, calls.count(("excel", "inventory-2030-08.xlsx")))
        self.assertEqual(2, calls.count(("backup", "inventory-2030-08-09.xlsx")))

    def test_archive_moves_each_source_once_to_its_configured_date_folder(self):
        with self.sessions.begin() as session:
            session.add_all((
                InventorySourceFileModel(
                    id="processed", tenant_id="t", external_source_id="src", drive_file_id="drive-a",
                    filename="a.jpg", drive_modified_time=__import__("datetime").datetime.now(__import__("datetime").timezone.utc),
                ),
                InventorySourceFileModel(
                    id="reupload", tenant_id="t", external_source_id="src", drive_file_id="drive-b",
                    filename="b.jpg", drive_modified_time=__import__("datetime").datetime.now(__import__("datetime").timezone.utc),
                ),
            ))
        moves = []
        class Drive:
            async def __aenter__(inner): return inner
            async def __aexit__(inner, *_): return None
            async def ensure_child_folder(inner, parent, name):
                return type("Folder", (), {"id": f"{parent}/{name}"})()
            async def move_file(inner, drive_file_id, destination):
                moves.append((drive_file_id, destination))
        service = InventoryExportService(
            self.sessions, token_resolver=self.service.token_resolver, client_factory=lambda _: Drive()
        )
        rows = [
            ("processed", "drive-a", {}, "finalized"),
            ("reupload", "drive-b", {}, "needs_reupload"),
        ]
        import asyncio
        asyncio.run(service._move_rows("connection", rows, {"finalized":"processed-root", "needs_reupload":"reupload-root"}, DAY, "t"))
        self.assertEqual([("drive-a", "processed-root/2030-08-09"), ("drive-b", "reupload-root/2030-08-09")], moves)
        with self.sessions() as session:
            retried = [
                (row.id, row.drive_file_id, row.provider_metadata_json, "finalized" if row.id == "processed" else "needs_reupload")
                for row in session.scalars(select(InventorySourceFileModel).order_by(InventorySourceFileModel.id))
            ]
        asyncio.run(service._move_rows("connection", retried, {"finalized":"processed-root", "needs_reupload":"reupload-root"}, DAY, "t"))
        self.assertEqual(2, len(moves))
    def test_archive_retry_reuses_completed_workbooks_without_upload(self):
        result = self.service.export("t", DAY)
        self.assertEqual(2, len(self.calls))
        with self.sessions.begin() as session:
            row = session.get(InventoryExportModel, result.id)
            row.archive_status = "retryable_failure"
            row.archive_error_code = "inventory_archive_drive_failure"
        calls = []
        def retry(tenant_id, export_id, business_date, source_id):
            calls.append((tenant_id, export_id, business_date, source_id))
            with self.sessions.begin() as session:
                row = session.get(InventoryExportModel, export_id)
                row.archive_status = "completed"
                row.archive_error_code = None
        self.service._archive_after_export = retry
        retried = self.service.export("t", DAY)
        self.assertEqual("completed", retried.archive_status)
        self.assertEqual(1, len(calls))
        self.assertEqual(2, len(self.calls), "archive retry must not upload main or backup")
        self.service.export("t", DAY)
        self.assertEqual(1, len(calls), "completed archive is idempotent")
    def test_unfinalized_day_rejected_without_upload(self):
        with self.assertRaisesRegex(InventoryExportFailure, "not_finalized"):
            self.service.export("t", date(2030, 8, 10))
        self.assertEqual([], self.calls)

    def test_sheet_four_fingerprint_covers_value_formula_merge_style_and_layout(self):
        workbook = Workbook()
        workbook.create_sheet("two")
        workbook.create_sheet("three")
        protected = workbook.create_sheet(PROTECTED_SHEET)
        protected["A1"] = "keep"
        protected["B1"] = "=1+1"
        protected.merge_cells("C1:D1")
        protected["E1"].font = Font(bold=True)
        protected.row_dimensions[1].height = 28
        protected.column_dimensions["A"].width = 31
        protected.freeze_panes = "A2"
        before = sheet_fingerprint(protected)
        for mutate in (
            lambda: protected.__setitem__("A1", "changed"),
            lambda: protected.__setitem__("B1", "=2+2"),
            lambda: protected.unmerge_cells("C1:D1"),
            lambda: setattr(protected["E1"], "font", Font(italic=True)),
            lambda: setattr(protected.row_dimensions[1], "height", 29),
        ):
            mutate()
            self.assertNotEqual(before, sheet_fingerprint(protected))
            # Rebuild so every mutation is checked independently.
            workbook = Workbook()
            workbook.create_sheet("two")
            workbook.create_sheet("three")
            protected = workbook.create_sheet(PROTECTED_SHEET)
            protected["A1"] = "keep"
            protected["B1"] = "=1+1"
            protected.merge_cells("C1:D1")
            protected["E1"].font = Font(bold=True)
            protected.row_dimensions[1].height = 28
            protected.column_dimensions["A"].width = 31
            protected.freeze_panes = "A2"
            before = sheet_fingerprint(protected)

    def test_sheet_four_mutation_fails_before_any_upload(self):
        service = InventoryExportService(
            self.sessions,
            token_resolver=self.service.token_resolver,
            client_factory=self.service.client_factory,
        )
        original = service._populate_workbook

        def corrupt(workbook, business_date, transactions):
            original(workbook, business_date, transactions)
            workbook[PROTECTED_SHEET]["A1"] = "unsafe mutation"

        service._populate_workbook = corrupt
        with self.assertRaisesRegex(InventoryExportFailure, "sheet4_invariant"):
            service.export("t", DAY)
        self.assertEqual([], self.calls)
        with self.sessions() as session:
            row = session.scalar(select(InventoryExportModel))
            self.assertEqual("failed", row.status)
            self.assertEqual("inventory_sheet4_invariant_failed", row.error_code)



    def test_needs_reupload_move_retries_only_its_configured_destination(self):
        from datetime import datetime, timezone
        import asyncio
        with self.sessions.begin() as session:
            session.add(InventorySourceFileModel(
                id="reupload-retry", tenant_id="t", external_source_id="src", drive_file_id="drive-reupload",
                filename="retry.jpg", drive_modified_time=datetime.now(timezone.utc),
            ))
        moves = []
        class Drive:
            attempts = 0
            async def __aenter__(inner): return inner
            async def __aexit__(inner, *_): return None
            async def ensure_child_folder(inner, parent, name):
                return type("Folder", (), {"id": f"{parent}/{name}"})()
            async def move_file(inner, file_id, destination):
                Drive.attempts += 1
                moves.append((file_id, destination))
                if Drive.attempts == 1:
                    raise RuntimeError("transient-drive-move-failure")
        service = InventoryExportService(self.sessions, token_resolver=self.service.token_resolver, client_factory=lambda _: Drive())
        rows = [("reupload-retry", "drive-reupload", {}, "needs_reupload")]
        with self.assertRaisesRegex(RuntimeError, "transient-drive-move-failure"):
            asyncio.run(service._move_rows("connection", rows, {"needs_reupload": "02_CAN_CHUP_LAI"}, DAY, "t"))
        with self.sessions() as session:
            self.assertEqual({}, session.get(InventorySourceFileModel, "reupload-retry").provider_metadata_json or {})
        asyncio.run(service._move_rows("connection", rows, {"needs_reupload": "02_CAN_CHUP_LAI"}, DAY, "t"))
        with self.sessions() as session:
            archive = session.get(InventorySourceFileModel, "reupload-retry").provider_metadata_json["inventory_archive"]
        self.assertEqual("02_CAN_CHUP_LAI/2030-08-09", archive["destination_folder_id"])
        self.assertEqual("needs_reupload", archive["document_status"])
        with self.sessions() as session:
            source = session.get(InventorySourceFileModel, "reupload-retry")
            persisted_rows = [(source.id, source.drive_file_id, source.provider_metadata_json, "needs_reupload")]
        asyncio.run(service._move_rows("connection", persisted_rows, {"needs_reupload": "02_CAN_CHUP_LAI"}, DAY, "t"))
        self.assertEqual(2, len(moves), "successful retry must make later movement idempotent")
        self.assertTrue(all(destination.startswith("02_CAN_CHUP_LAI/") for _, destination in moves))

    def test_drive_failure_logs_never_include_credentials(self):
        access, refresh, authorization = "access-token-abc", "refresh-token-def", "Authorization: Bearer access-token-abc"
        class Drive:
            async def __aenter__(inner): return inner
            async def __aexit__(inner, *_): return None
            async def upload_file(inner, *_):
                raise RuntimeError(f"provider failed {access} {refresh} {authorization}")
        service = InventoryExportService(self.sessions, token_resolver=lambda _: access, client_factory=lambda _: Drive())
        with self.assertLogs("app.modules.inventory.exports.service", "WARNING") as logs:
            with self.assertRaisesRegex(InventoryExportFailure, "inventory_export_drive_failure"):
                service.export("t", DAY)
        rendered = "\\n".join(logs.output)
        self.assertIn("inventory_export_upload_failed", rendered)
        self.assertNotIn(access, rendered)
        self.assertNotIn(refresh, rendered)
        self.assertNotIn("Authorization", rendered)
        with self.sessions() as session:
            row = session.scalar(select(InventoryExportModel))
            self.assertEqual("inventory_export_drive_failure", row.error_code)
            self.assertNotIn(access, row.error_message or "")
            self.assertNotIn(refresh, row.error_message or "")

if __name__ == "__main__":
    unittest.main()